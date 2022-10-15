from openpyxl import load_workbook, Workbook
# from MyWeb import settings
# import os


def read_all_data(path):
    excel = load_workbook(path)
    data = {}
    sheetnames = excel.sheetnames
    for sheet in excel:
        max_row = sheet.max_row
        max_col = sheet.max_column
        rows = handle_rows_for_previw(sheet, max_row, max_col)
        data[sheet.title] = rows
    # print(data)
    excel.close()
    return sheetnames, data


def handle_rows_for_previw(sheet, max_row, max_col):
    rows = []
    for i in range(max_row):
        row = []
        for j in range(max_col):
            row.append(sheet.cell(i + 1, j + 1).value)
        rows.append(row)
    return rows


def read_data_by_name(path, mth, sheetname=None):
    excel = load_workbook(path)
    data, sheet = None, None
    if not sheetname:
        sheet = excel.worksheets[0]
    else:
        if sheetname in excel.sheetnames:
            sheet = excel[sheetname]
    if sheet:
        max_row = sheet.max_row
        max_col = sheet.max_column
        data = mth(sheet, max_row, max_col)
    excel.close()
    return data


def handle_rows_for_api_ds(sheet, max_row, max_col):
    rows = []
    for i in range(1, max_row):
        if not sheet.cell(i + 1, 2).value:
            continue
        row = {'desc': sheet.cell(i + 1, 1).value, 'p_name': sheet.cell(i + 1, 2).value}
        values = []
        for j in range(2, max_col):
            values.append(sheet.cell(i + 1, j + 1).value)
        row['values'] = values
        rows.append(row)
    return rows



